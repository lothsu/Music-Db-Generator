import os
import pathlib
from openpyxl import Workbook

acceptedFolders = ['flac', 'low quality', 'other']
acceptedFiles = [".mp3" , ".opus", ".flac", ".m4a"]

directory = pathlib.Path(__file__).parent.resolve()

musicCollection = os.scandir(directory)

wb = Workbook()
ws = wb.active
ws.title = "Übersicht"



def makeHeaders():
    ws['A1'] = "Künstler"
    ws['B1'] = "Album"
    ws['C1'] = "Titel"
    ws['D1'] = "Qualität"
    ws['E1'] = "Format"


def fillInXl():
    print("Files and Directories in '% s':" % directory)
    rowCounter = 2

    for quality in musicCollection:
        if quality.is_dir() and quality.name in acceptedFolders:
            print(quality.name)
            artists = os.scandir(quality)
            
            for artist in artists:
                if artist.is_dir():
                    #print(artist.name)
                    albums = os.scandir(artist)
                    
                    for album in albums:
                        if album.is_dir():
                            #print(album.name)
                            titels = os.scandir(album)
                            
                            for titel in titels:
                                fileName = os.path.splitext(titel.name)
                                if titel.is_file() and fileName[1] in acceptedFiles:
                                    print(fileName[0])
                                
                                    ws.cell(row = rowCounter, column = 1).value = artist.name
                                    ws.cell(row = rowCounter, column = 2).value = album.name
                                    ws.cell(row = rowCounter, column = 3).value = fileName[0]
                                    ws.cell(row = rowCounter, column = 4).value = quality.name
                                    ws.cell(row = rowCounter, column = 5).value = fileName[1]

                                    rowCounter += 1


def saveXl():
     wb.save('Musik DB.xlsx')


makeHeaders()
fillInXl()
saveXl()
